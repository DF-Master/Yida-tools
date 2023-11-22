from Bio import PDB


def cal_distance_pos(animo_pos1,
                     animo_pos2,
                     pdb,
                     name='default',
                     chain1='A',
                     animo_core1='CA',
                     animo_core2='CA',
                     in_one_domain=True,
                     autoprint=False):
    parser = PDB.PDBParser()
    structure = parser.get_structure(name, pdb)
    chain1 = chain1
    model1 = structure[0][chain1]
    try:
        residue1 = model1[animo_pos1]  # Start from 1 not 0
        residue2 = model1[animo_pos2]
        atom1 = residue1[animo_core1]
        atom2 = residue2[animo_core2]
        distance = atom1 - atom2
        if autoprint == True:
            print('Distance between', residue1.get_resname(), animo_core1,
                  'and', residue2.get_resname(), animo_core2, 'is ', distance)
        return distance
    except:
        print(
            animo_pos1,
            animo_pos2,
            pdb,
        )
        return None


import openpyxl
# Define the path to the Excel file
# file_path = r"G:/MSdata/220329ADK/adk/spotlink/data_analyser.xlsx"
file_path = r"G:\MSdata\202310AHL\ADRM1\reports\cross-link\duplicates\report/ADRM1_residue_counts_dts_1_rts_1_type2.xlsx"
pdb_file = r"G:/MSdata/pdb/ADRM1.pdb"

# Load the workbook
workbook = openpyxl.load_workbook(file_path)

# Select the sheets to compare
sheet_names = workbook.sheetnames

# Colume define
column1 = 25  # Column Y,start from 1
column2 = 29  # Column AC

# Distance Column
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]

    max_row = sheet.max_row
    max_col = sheet.max_column

    if max_col > column1 and max_col > column2:
        # Add header for new column
        sheet.cell(row=1, column=max_col + 1, value=pdb_file.split("/")[-1])

        # Iterate through rows to calculate the distance and update the new column
        for row in range(2, max_row + 1):
            pos1 = sheet.cell(row=row, column=column1).value
            pos2 = sheet.cell(row=row, column=column2).value
            distance = cal_distance_pos(pos1, pos2, pdb_file, autoprint=True)
            sheet.cell(row=row, column=max_col + 1, value=distance)

# Save the modified workbook
workbook.save(file_path)