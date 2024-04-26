from openpyxl import load_workbook

# Define the file path
file_dir = r'C:\Users\jiang\OneDrive\Research\tc\Data\2024\03'
file_name = r'41467_2020_20159_MOESM4_ESM_PXD019437.xlsx'
# file_dir = r"G:\MSdata\220329ADK\adk\plink\pLink_task_2022.04.02.20.19.24\reports"
# file_name = "data_analyser_plink_240301.xlsx"
file_path = file_dir + '\\' + file_name
sheet_name = 'inter_protein_HRST_Others'

# Load the workbook
workbook = load_workbook(file_path)

# Select the worksheet
worksheet = workbook[sheet_name]

# Define Pos Column !!!!!
# column1 = 24  # Column Y,start from 0
# column2 = 28  # Column AC,start from 0
column1 = 1
column2 = 4
dis_column = 25
# chain_column = 0
# dimer_column = 18
protein_name = '6zfb_dimeric_RNAP-_-HelD'
chain_dic = {"RpoA": "U", "RpoB": "X", "RpoC": "Y", "HelD": "H", "RpoE": "E"}

# Loop through each row
for row in worksheet.iter_rows(min_row=1, values_only=True):
    # Get the values from Column X and Column AD
    value1 = row[column1]
    value2 = row[column2]
    dis = row[dis_column]
    # chain_name = row[chain_column]
    # dimer = row[dimer_column]
    try:
        dis = float(dis)
        if dis < 20:
            color = "yellow"
        else:
            color = "red"
        if row[18] == 1:
            chain_dic_use = {
                "RpoA": "V",
                "RpoB": "X",
                "RpoC": "Y",
                "HelD": "H",
                "RpoE": "E"
            }
        else:
            chain_dic_use = chain_dic
        chain1 = chain_dic_use[row[0]]
        chain2 = chain_dic_use[row[3]] if row[17] == 0 else chain_dic_use[
            row[3]].lower()
        print(
            f"distance {row[24]}{value1}-{value2}, /{protein_name}//{chain1}/{value1}/CA, /{protein_name}//{chain2}/{value2}/CA; color {color}, {row[24]}{value1}-{value2};"
        )
    except:
        print(dis, ": Invalid input. Not a number.")
