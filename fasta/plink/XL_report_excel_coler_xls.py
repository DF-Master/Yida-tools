import openpyxl
from openpyxl.styles import PatternFill

# Define the path to the Excel file
file_path = r"C:\Users\jiang\OneDrive\Research\tc\articles\diazirine\crosslink\BSA\reports-3\BSA-type2XL\combined_data\duplicates\report\repeat_experiment_intersection\LF_residue_counts_dts_3_rts_3.xlsx"
# file_path = r"G:/MSdata/220329ADK/adk/plink/pLink_task_2022.04.02.20.19.24/reports/data_analyser_plink.xlsx"

# Load the workbook
workbook = openpyxl.load_workbook(file_path)

# Select the sheets to compare
sheet1 = workbook["200"]
sheet2 = workbook["1"]

# Colume define
# column1 = 24  # Column Y,start from 0
# column2 = 28  # Column AC

column1 = 26  # Column AA,start from 0
column2 = 32  # Column AG

# Identify error
error_concept = 2


# def color_change():
def color_change(color="00FF00", fill_type="solid"):
    return PatternFill(start_color=color, fill_type=fill_type)


# Iterate through each row in the first sheet
for row1 in sheet1.iter_rows(min_row=2):
    # Get the volume values from column O and S
    volume1 = row1[column1]
    volume2 = row1[column2]

    # Iterate through each row in the second sheet
    for row2 in sheet2.iter_rows(min_row=2):
        # Get the volume values from column O and S
        volume3 = row2[column1]
        volume4 = row2[column2]
        # Compare the volume values and change background color if they match
        if volume1.value == volume3.value and volume2.value == volume4.value:
            volume1.fill = color_change()
            volume2.fill = color_change()
            volume3.fill = color_change()
            volume4.fill = color_change()
            break  # Stop iterating through the second sheet if a match is found
        elif abs(volume1.value - volume3.value) <= error_concept and abs(
                volume2.value - volume4.value) <= error_concept:
            volume1.fill = color_change(color="FFFF00")
            volume2.fill = color_change(color="FFFF00")
            volume3.fill = color_change(color="FFFF00")
            volume4.fill = color_change(color="FFFF00")
        elif abs(volume1.value - volume2.value) <= 10:
            volume1.fill = color_change(color="888888")
            volume2.fill = color_change(color="888888")
        elif abs(volume3.value - volume4.value) <= 10:
            volume3.fill = color_change(color="888888")
            volume4.fill = color_change(color="888888")

    # Save the modified workbook
    workbook.save(file_path)