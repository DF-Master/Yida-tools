import os
import openpyxl as xl

# Path to the folder containing the xlsx files
folder_path = r"C:\Users\jiang\OneDrive\Research\tc\articles\diazirine\crosslink\BSA\reports-3\BSA-type2XL\combined_data\duplicates"

# List of valid characters for columns X and AD
valid_characters = "ACDEFGHIKLMNPQRSTVWY"

# Create a new xlsx file for the report
report_folder = os.path.join(folder_path, "report")
os.makedirs(report_folder, exist_ok=True)
report_filename = os.path.join(report_folder,
                               "BSA_residue_counts_dts_3_rts_2.xlsx"
                               )  # Change the name of the report file !!!!!
report_workbook = xl.Workbook()
report_workbook.remove(report_workbook.active)
# Create a new sheet in the workbook

# Define Pos Column !!!!!
column1 = 23  # Column X,start from 0
colume2 = 29  # Column AD,start from 0
duplicate_ts = 3  # Least duplicate times
repeat_ts = 2  # Least repeat times
score_ts = 1  # Score threshold
repeat_ts_column = 38  # Column AM,start from 0
duplicate_ts_column = 39  # Column AN,start from 0
scorecolume = 10  # Column K,start from 0
report_sheet = report_workbook.create_sheet("Residue Counts")

# Iterate through each xlsx file in the folder
for file_name in os.listdir(folder_path):
    power = file_name.split(".")[0]
    report_sheet = report_workbook.create_sheet(power)

    if file_name.endswith(".xlsx"):
        file_path = os.path.join(folder_path, file_name)

        # Load the xlsx file
        workbook = xl.load_workbook(file_path)
        sheet = workbook.active

        # Create a dictionary to store word counts for each residue
        word_counts = {residue: 0 for residue in valid_characters}

        # Count the repeat times of each word other than "K" in columns
        for row in sheet.iter_rows(min_row=1, values_only=True):
            cell_x = row[column1]
            cell_ad = row[colume2]
            cell_ts = row[repeat_ts_column]
            cell_score = row[scorecolume]
            cell_duplicate_ts = row[duplicate_ts_column]
            if cell_x is not None and cell_x != "K" and cell_x in valid_characters and cell_ts >= repeat_ts and cell_score <= score_ts and cell_duplicate_ts >= duplicate_ts:
                if cell_x in word_counts:
                    word_counts[cell_x] += 1
                    # add rows into sheet named "file_name"
                    report_sheet.append(row)
            elif cell_ad is not None and cell_ad in valid_characters and cell_ts >= repeat_ts and cell_score <= score_ts and cell_duplicate_ts >= duplicate_ts:
                if cell_ad in word_counts:
                    word_counts[cell_ad] += 1
                    # add rows into sheet named "file_name"
                    report_sheet.append(row)
            elif report_sheet.max_row == 1 and report_sheet.max_column == 1:
                report_sheet.append(row)
            else:
                # print(cell_x, cell_ad)
                pass

        print(word_counts)
        # Write the word counts to the report file

        if file_name != "word_counts.xlsx":
            # Get the active sheet of the report workbook
            report_sheet = report_workbook["Residue Counts"]

            # Add file name as header column
            report_sheet.cell(
                row=1, column=report_sheet.max_column +
                1).value = file_name.split("_")[-1].split(".")[0] + "mW"

            # Add word and count for each file in new columns
            column_num = report_sheet.max_column  # Get the column number for the file name
            start_row = 2  # Start from the second row, as first row contains headers
            for word, count in word_counts.items():
                report_sheet.cell(row=start_row,
                                  column=column_num).value = word
                report_sheet.cell(row=start_row,
                                  column=column_num + 1).value = count
                start_row += 1

# Save the modified report workbook
report_workbook.save("word_counts.xlsx")

# Save and close the report file
report_workbook.save(report_filename)
report_workbook.close()